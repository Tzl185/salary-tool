import streamlit as st
import os
import pandas as pd
from openpyxl import load_workbook
import tempfile
import zipfile

# 原封不动移植的函数
def process_file_a(folder_path, output_file="文件A_汇总结果.xlsx"):
    """
    处理文件夹中的Excel表，生成汇总的文件A
    返回文件A的路径和所有数值字典
    """
    all_data = []
    all_values = {}
    
    for filename in os.listdir(folder_path):
        if filename.endswith(('.xls', '.xlsx')) and not filename.startswith('~$'):
            filepath = os.path.join(folder_path, filename)
            try:
                # 读取Excel，第四行作为表头
                df = pd.read_excel(filepath, header=3)
                print(f"处理文件: {filename}")
                
                # 预算单位列（B列，索引1）
                budget_unit_col = df.columns[1]
                
                # 工资类型列，Q到AD列（索引16到29）
                wage_cols = df.columns[16:30]
                
                # 选取需要的列
                df_filtered = df[[budget_unit_col] + list(wage_cols)]
                
                # 将工资列转换为数值型，非数字转NaN，再用0填充
                df_filtered[wage_cols] = df_filtered[wage_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
                
                # 按预算单位分组求和
                df_grouped = df_filtered.groupby(budget_unit_col).sum()
                
                # 收集所有数值
                for budget_unit, row in df_grouped.iterrows():
                    for wage_type in wage_cols:
                        value = row[wage_type]
                        print(f"原始工资类型: {wage_type}, 值: {value}")  # 调试用
                        original_wage_type = wage_type  # 保存原始值
                        wage_type = wage_type.strip()  # 去除前后空格
                        if "绩效工资" in wage_type:
                            wage_type = wage_type.replace("绩效工资", "基础性绩效")
                        if "行政医疗" in wage_type:
                            wage_type = wage_type.replace("行政医疗", "职工基本医疗（行政）")
                        elif "事业医疗" in wage_type:  # 使用elif避免重复替换
                            wage_type = wage_type.replace("事业医疗", "基本医疗（事业）")
                        elif "医疗保险" in wage_type:  # 可能还有其他表述
                            wage_type = wage_type.replace("医疗保险", "基本医疗")
                        print(f"处理后工资类型: {wage_type}")  # 调试用
                        key = (str(budget_unit).strip(), str(wage_type).strip())
                        all_values[key] = value
                        if "医疗" in wage_type:
                            print(f"医疗数值记录 - 单位: {budget_unit}, 类型: {wage_type}, 值: {value}")

                if df_grouped is not None and not df_grouped.empty:
                    all_data.append(df_grouped)
                
            except Exception as e:
                print(f"处理文件 {filename} 出错: {e}")
    
    if all_data:
        # 合并所有文件的汇总结果
        df_all = pd.concat(all_data)
        # 按预算单位再次汇总
        df_final = df_all.groupby(df_all.index).sum()
        
        output_path = os.path.join(folder_path, output_file)
        df_final.to_excel(output_path)
        print(f"\n汇总结果已保存到: {output_path}")
        
        print(f"\n总共收集到 {len(all_values)} 个数值")
        return output_path, all_values
    else:
        print("没有找到有效数据")
        return None, None

def update_file_b(file_a_path, file_b_path):
    """
    用文件A中的所有数值更新文件B的J列，保留原有格式
    """
    try:
        # 1. 从文件A中读取数据
        df_a = pd.read_excel(file_a_path, index_col=0)
        wage_cols = df_a.columns
        
        # 提取所有数值
        all_values = {}
        for budget_unit, row in df_a.iterrows():
            for wage_type in wage_cols:
                value = row[wage_type]
                if "绩效工资" in wage_type:
                    wage_type = wage_type.replace("绩效工资", "基础性绩效")
                if "行政医疗" in wage_type:
                    wage_type = wage_type.replace("行政医疗", "职工基本医疗（行政）")
                elif "事业医疗" in wage_type:  # 使用elif避免重复替换
                    wage_type = wage_type.replace("事业医疗", "基本医疗（事业）")
                elif "医疗保险" in wage_type:  # 可能还有其他表述
                    wage_type = wage_type.replace("医疗保险", "基本医疗")
                key = (str(budget_unit).strip(), str(wage_type).strip())
                all_values[key] = value

        # 2. 使用openpyxl直接操作Excel文件
        wb = load_workbook(file_b_path)
        sheet = wb.active
        
        # J列的索引（从1开始计数）
        j_col_index = 10
        
        # 3. 更新J列数据
        match_count = 0
        for row_idx in range(2, sheet.max_row + 1):  # 从第2行开始
            unit_cell = sheet.cell(row=row_idx, column=1)
            unit_info = str(unit_cell.value).strip() if unit_cell.value else ""
            
            budget_cell = sheet.cell(row=row_idx, column=2)
            budget_project = str(budget_cell.value).strip() if budget_cell.value else ""
            
            # 清理单位信息
            unit_info_cleaned = unit_info.replace("-", "").replace(" ", "")
            
            # 查找匹配
            matched = False
            for (budget_unit, wage_type), value in all_values.items():
                budget_unit_cleaned = budget_unit.replace("-", "").replace(" ", "")
                
                # 匹配条件
                unit_match = (budget_unit_cleaned in unit_info_cleaned) or (unit_info_cleaned in budget_unit_cleaned)
                wage_match = wage_type in budget_project
                
                if unit_match and wage_match:
                    # 更新单元格值，保留原有样式
                    sheet.cell(row=row_idx, column=j_col_index).value = value
                    match_count += 1
                    matched = True
                    print(f"匹配成功: 行{row_idx} 单位:'{unit_info}'⊇'{budget_unit}', 项目:'{budget_project}'⊇'{wage_type}', 值:{value}")
                    break
            
            if not matched and row_idx < 7:  # 打印前5行未匹配情况
                print(f"未匹配: 行{row_idx} 单位:'{unit_info}', 项目:'{budget_project}'")
        
        # 4. 保存更新后的文件B
        output_path = os.path.join(os.path.dirname(file_b_path), "updated_" + os.path.basename(file_b_path))
        wb.save(output_path)
        print(f"\n总共完成 {match_count} 处匹配")
        print(f"已保存更新后的文件B到: {output_path}")
        return output_path
    
    except Exception as e:
        print(f"\n更新文件B出错: {e}")
        return None

def process_and_download(uploaded_zip, uploaded_template):
    """处理上传的文件并提供下载"""
    with tempfile.TemporaryDirectory() as temp_dir:
        try:
            # 保存ZIP文件
            zip_path = os.path.join(temp_dir, "upload.zip")
            with open(zip_path, "wb") as f:
                f.write(uploaded_zip.getvalue())
            
            # 解压ZIP
            extract_dir = os.path.join(temp_dir, "extracted")
            os.makedirs(extract_dir, exist_ok=True)
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(extract_dir)
            
            # 保存模板文件
            template_path = os.path.join(temp_dir, "template.xlsx")
            with open(template_path, "wb") as f:
                f.write(uploaded_template.getvalue())
            
            # 调用原有处理逻辑
            file_a_path, _ = process_file_a(extract_dir)
            if file_a_path:
                result_path = update_file_b(file_a_path, template_path)
                if os.path.exists(result_path):
                    with open(result_path, "rb") as f:
                        st.download_button(
                            "下载处理结果",
                            data=f,
                            file_name="工资调整结果.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    st.success("处理完成！")
                else:
                    st.error("生成结果文件失败")
            else:
                st.error("ZIP压缩包内未找到有效数据")
        except Exception as e:
            st.error(f"处理过程中出错: {str(e)}")

# Streamlit界面
def main():
    st.title("工资调整自动汇总与填充工具")
    
    # 文件上传区域
    col1, col2 = st.columns(2)
    with col1:
        uploaded_zip = st.file_uploader("上传ZIP压缩包", type="zip")
    with col2:
        uploaded_template = st.file_uploader("上传模板文件", type=["xlsx"])
    
    # 添加明确的处理按钮
    if uploaded_zip and uploaded_template:
        if st.button("开始处理", type="primary", help="点击后开始处理数据"):
            with st.spinner("正在处理，请稍候..."):
                process_and_download(uploaded_zip, uploaded_template)
    else:
        st.warning("请先上传ZIP压缩包和模板文件")
        
if __name__ == "__main__":
    main()
